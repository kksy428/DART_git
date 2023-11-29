import openpyxl
from collections import Counter

# 워크북(workbook) 불러오기
workbook = openpyxl.load_workbook("students.xlsx")

# 활성화된 워크시트(worksheet) 가져오기
worksheet = workbook.active

# 학과 정보가 담길 리스트
departs = []

# 각 행(row)을 하나씩 확인하며
max_row = worksheet.max_row
for row in range(1, max_row + 1):
    # 학과 문자열을 확인하여 리스트에 담기
    value = worksheet.cell(row=row, column=6).value
    departs.append(value)

# 각 학과의 출현 빈도 계산하기
counter = Counter(departs)
print("출현 빈도:", counter)

# 각 학과별 사람 수 기록하기
worksheet.cell(row=1, column=9).value = "학과"
worksheet.cell(row=1, column=10).value = "사람 수"
row = 2
for depart in counter:
    worksheet.cell(row=row, column=9).value = depart
    worksheet.cell(row=row, column=10).value = counter[depart]
    row += 1

# 막대 차트(Bar Chart) 만들기
bar_chart = openpyxl.chart.BarChart()
bar_chart.title = "학과 통계"
bar_chart.x_axis.title = "학과"
bar_chart.y_axis.title = "사람 수"

"""
데이터의 타이틀이 포함되어 있다면, titles_from_data=True
데이터가 각 행에 걸쳐서(행우선) 차례대로 기입되어 있다면, from_rows=True
"""
# 데이터를 이용해 차트에 그리기
datas = openpyxl.chart.Reference(worksheet, min_row=2, min_col=9, max_row=2 + len(counter) - 1, max_col=10)
bar_chart.add_data(datas, from_rows=True, titles_from_data=True)

# 만들어진 차트(chart)를 "A32"의 위치에 그리기
worksheet.add_chart(bar_chart, "A32")
# 완성된 엑셀 파일 저장하기
workbook.save("bar_chart.xlsx")

import openpyxl
from collections import Counter

# 워크북(workbook) 불러오기
workbook = openpyxl.load_workbook("students.xlsx")

# 활성화된 워크시트(worksheet) 가져오기
worksheet = workbook.active

# 학년 정보가 담길 리스트
grades = []

# 각 행(row)을 하나씩 확인하며
max_row = worksheet.max_row
for row in range(1, max_row + 1):
    # 학년 정보를 확인하여 리스트에 담기
    value = worksheet.cell(row=row, column=7).value
    grades.append(value)

# 각 학년의 출현 빈도 계산하기
counter = Counter(grades)
print("출현 빈도:", counter)

# 각 학년별 사람 수 기록하기
worksheet.cell(row=1, column=9).value = "학년"
worksheet.cell(row=1, column=10).value = "사람 수"
row = 2
for grade in counter:
    worksheet.cell(row=row, column=9).value = grade
    worksheet.cell(row=row, column=10).value = counter[grade]
    row += 1

# 파이 차트(Pie Chart) 만들기
pie_chart = openpyxl.chart.PieChart()
pie_chart.title = "학년 통계"

"""
데이터의 타이틀이 포함되어 있다면, titles_from_data=True
데이터가 각 행에 걸쳐서(행우선) 차례대로 기입되어 있다면, from_rows=True
"""
# 불러올 데이터 명시
labels = openpyxl.chart.Reference(worksheet, min_row=2, min_col=9, max_row=2 + len(counter) - 1, max_col=9)
datas = openpyxl.chart.Reference(worksheet, min_row=1, min_col=10, max_row=2 + len(counter) - 1, max_col=10)

pie_chart.add_data(datas, from_rows=False, titles_from_data=True) # 차트에 데이터(data) 추가
pie_chart.set_categories(labels) # 차트에 레이블(label) 추가

# 만들어진 차트(chart)를 "A32"의 위치에 그리기
worksheet.add_chart(pie_chart, "A32")
# 완성된 엑셀 파일 저장하기
workbook.save("pie_chart.xlsx")

import openpyxl

workbook = openpyxl.Workbook() # 워크북(workbook) 생성

# 활성화된 워크시트(worksheet) 가져오기
worksheet = workbook.active

# 특정한 파일 경로에서 이미지 불러오기
image = openpyxl.drawing.image.Image("python.png")
# 이미지를 "B2"의 위치에 그리기
worksheet.add_image(image, "B2")

# 완성된 엑셀 파일 저장하기
workbook.save("image.xlsx")

