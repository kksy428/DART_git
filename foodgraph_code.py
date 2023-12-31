import openpyxl
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

workbook = openpyxl.load_workbook("역이름.xlsx")
worksheet = workbook.active
food = []
max_row = worksheet.max_row
for row in range(1,max_row + 1):
    value = worksheet.cell(row=row, column=1).value
    food.append(value)


def topfood(text):
    re_text = 'https://search.daum.net/search?w=tot&q='+text
    request = requests.get(re_text)
    html = request.text.strip()
    soup = BeautifulSoup(html, 'html.parser')

    data1 = soup.find_all('a', class_ ='fn_tit')
    data2 = soup.find_all('span', class_ ='f_red')

    tit = []
    star = []

    for i,v in enumerate(data1):
        tit.append(v.get_text())

    for i,v in enumerate(data2):
        Test = v.get_text()
    
        if '.' in Test :
            star.append(float(v.get_text()))
    


    workbook = Workbook() 
    worksheet = workbook.active
    worksheet.append(['가게 이름','별점'])
    for t,s in zip(tit,star):
        worksheet.append([t,s])
    
    bar_chart = openpyxl.chart.BarChart()
    bar_chart.width = 17
    bar_chart.height = 14

    bar_chart.title = f"{text}"
    bar_chart.x_axis.title = "가게 이름"
    bar_chart.y_axis.title = "별점"

    datas = openpyxl.chart.Reference(worksheet, min_row=2, min_col=1, max_row=7, max_col=2)
    bar_chart.add_data(datas, from_rows=True, titles_from_data=True)
    worksheet.add_chart(bar_chart, "A9")

    workbook.save("food_list.xlsx")



while (1):
    name = input('7호선 중 원히는 역 이름을 검색하세요. (역 뻬고)')

    if name in food :
        print(f'7호선 {name}역 맛집 분석 시작하겠습니다.')
        topfood(f'{name}역맛집')
        break
    else:
        print(f'{name}역은 7호선에 존재하기 않습니다.')




